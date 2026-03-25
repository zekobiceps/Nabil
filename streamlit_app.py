import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import os
from pathlib import Path
from openpyxl import load_workbook
import io as _io

# ═══════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Générateur — Fin Période d'Essai",
    page_icon="📋",
    layout="wide",
)

# ═══════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ═══════════════════════════════════════════════════════════

def parse_date(v):
    """Parse une valeur date de façon sécurisée."""
    if v is None:
        return None
    try:
        if isinstance(v, float) and pd.isna(v):
            return None
    except TypeError:
        pass
    try:
        return pd.to_datetime(v)
    except (ValueError, TypeError):
        return None


def get_gender_info(civ, default="Mme"):
    """Retourne (titre, is_female) selon la civilité."""
    if not civ or (isinstance(civ, float) and pd.isna(civ)):
        civ = default
    s = str(civ).upper().strip()
    if s in ("M.", "MR", "M", "MONSIEUR") or s.startswith("MR ") or s.startswith("M. "):
        return "M.", False
    if "MLLE" in s or "MADEMOISELLE" in s:
        return "Mlle", True
    return "Mme", True  # Mme par défaut


def build_titularisation(nom, prenom, poste, date_entree, date_fin,
                          titre="Mme", is_female=True):
    """Génère le message de titularisation."""
    verb = "recrutée" if is_female else "recruté"
    return (
        "Bonjour ,\n\n"
        f"Je me permets de vous contacter au sujet de la titularisation de "
        f"{titre} {nom} {prenom}, {verb} en qualité de {poste} "
        f"depuis le {date_entree.strftime('%d/%m/%Y')}. "
        f"Sa période d'essai arrive à son terme le {date_fin.strftime('%d/%m/%Y')}.\n\n"
        "De ce fait, je vous prie de bien vouloir remplir le document ci-joint "
        "afin de confirmer ou non sa titularisation.\n\n"
        "Sincères salutations,"
    )


def build_prolongement(individu, nom, prenom, poste, direction, sup,
                        date_entree, date_fin_1ere, titre="Mme", is_female=True):
    """Génère le message de prolongement de période d'essai."""
    collab = "collaboratrice" if is_female else "collaborateur"
    hdr = "\t".join([
        titre, "NOM", "PRENOM", "FONCTION",
        "CHANTIER CTRL PRES", "SUP ", "DATE D'EMBAUCHE",
        "DATE FIN  1ERE PERIODE D'ESSAI"
    ])
    data_row = "\t".join([
        str(individu), nom, prenom, poste, direction, sup,
        date_entree.strftime("%Y-%m-%d"),
        date_fin_1ere.strftime("%d/%m/%Y"),
    ])
    return (
        "Bonjour , \n\n"
        f"Nous vous informons que la {collab} mentionné(e) ci-dessous "
        "arrive à la fin de leur première période d'essai.\n\n"
        f"{hdr}\n{data_row}\n\n"
        "Pouvez-vous nous confirmer si vous les considérez aptes à bénéficier "
        "d'une prolongation de leur période d'essai ?\n"
        "Nous vous prions de bien vouloir nous répondre par retour de mail.\n\n"
        "Cordialement, "
    )


def auto_map_columns(columns):
    """Détection automatique de la correspondance des colonnes."""
    mapping = {}
    for col in columns:
        cu = col.upper().replace(" ", "").replace("(", "").replace(")", "").strip()
        if cu == "NOM":
            mapping.setdefault("NOM", col)
        elif cu == "PRENOM":
            mapping.setdefault("PRENOM", col)
        elif cu in ("LIB", "LIBPOSTE", "POSTE", "FONCTION", "LIBELLEPOSTE"):
            mapping.setdefault("LIB", col)
        elif cu in ("LIB80", "LIB80DIRECTION", "DIRECTION", "CHANTIER", "CHANTIERCTRLPRES"):
            mapping.setdefault("LIB80", col)
        elif cu in ("SUP", "NOMDURESPONSABLE", "NOMDURESPHIERARCHIQUE", "NOMRESPONSABLE"):
            mapping.setdefault("SUP", col)
        elif "DATE" in cu and ("ENTREE" in cu or "EMBAUCHE" in cu):
            mapping.setdefault("DATE_ENTREE", col)
        elif ("RENOUVEL" in cu and "DATE" in cu) or cu in ("RENOUVELLEMENTDATE", "DATERENOUVELLEMENT"):
            mapping.setdefault("DATE_RENOUVELLEMENT", col)
        elif cu in ("INDIVIDU", "MATRICULE", "MAT", "MLE", "MLLE"):
            mapping.setdefault("INDIVIDU", col)
        elif cu in ("CIVILITE", "TITRE", "CIVIL", "CIVILITÉ"):
            mapping.setdefault("CIVILITE", col)
        elif "EMAIL" in cu or "MAIL" in cu:
            mapping.setdefault("EMAIL", col)
    return mapping


def get_safe_str(row, col):
    """Retourne une chaîne propre depuis une cellule du DataFrame."""
    if col is None:
        return ""
    v = row.get(col, "")
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def resolve_attachment_path(exact_name, name_prefix=None):
    """Résout un chemin de pièce jointe depuis le dossier de l'application."""
    app_dir = Path(__file__).resolve().parent
    exact = app_dir / exact_name
    if exact.exists():
        return str(exact)
    if name_prefix:
        for candidate in app_dir.glob(f"{name_prefix}*"):
            if candidate.is_file():
                return str(candidate)
    return str(exact)


# ═══════════════════════════════════════════════════════════
# SIDEBAR — CONFIGURATION
# ═══════════════════════════════════════════════════════════
with st.sidebar:
    st.title("⚙️ Configuration")

    msg_type = st.radio(
        "Type de message",
        ["📄 Titularisation", "🔄 Prolongement Période d'Essai"],
        index=0,
    )
    is_titularisation = msg_type.startswith("📄")

    st.caption("La durée d'essai est calculée automatiquement depuis DATE ENTREE et Renouvellement Date.")

    st.divider()
    # Civilité supprimée — ne plus demander la civilité par défaut

# ═══════════════════════════════════════════════════════════
# PAGE PRINCIPALE
# ═══════════════════════════════════════════════════════════
st.title("📋 Générateur — Fin de Période d'Essai")
mode_label = "Titularisation" if is_titularisation else "Prolongement Période d'Essai"
st.caption(f"Mode actif : **{mode_label}**")
st.divider()

# ── ÉTAPE 1 : Import ────────────────────────────────────────
st.subheader("① Importer la liste des collaborateurs")
st.caption("Colonnes attendues : NOM · PRENOM · LIB/POSTE · DATE ENTREE · Renouvellement Date (+ optionnel : LIB80/DIRECTION · SUP · INDIVIDU · EMAIL)")

uploaded = st.file_uploader(
    "Charger le fichier Excel ou CSV",
    type=["xlsx", "xls", "csv"],
    help="Format supporté : .xlsx, .xls, .csv"
)

if uploaded:
    try:
        if uploaded.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded)
        else:
            df = pd.read_excel(uploaded, engine="openpyxl" if uploaded.name.endswith("xlsx") else None)
        df = df.dropna(how="all").reset_index(drop=True)
    except Exception as e:
        st.error(f"❌ Impossible de lire le fichier : {e}")
        st.stop()

    st.success(f"✅ {len(df)} collaborateur(s) chargé(s)")

    with st.expander("📊 Aperçu du fichier importé", expanded=False):
        st.dataframe(df, use_container_width=True)

    detected = auto_map_columns(df.columns.tolist())
    col_nom = detected.get("NOM")
    col_prenom = detected.get("PRENOM")
    col_lib = detected.get("LIB")
    col_lib80 = detected.get("LIB80")
    col_sup = detected.get("SUP")
    col_date = detected.get("DATE_ENTREE")
    col_date_renouv = detected.get("DATE_RENOUVELLEMENT")
    col_individu = detected.get("INDIVIDU")
    # civilité supprimée — on ne détecte plus la colonne CIVILITE

    missing_required = []
    if not col_nom:
        missing_required.append("NOM")
    if not col_prenom:
        missing_required.append("PRENOM")
    if not col_lib:
        missing_required.append("LIB/POSTE")
    if not col_date:
        missing_required.append("DATE ENTREE")
    if not col_date_renouv:
        missing_required.append("Renouvellement Date")

    required_ok = len(missing_required) == 0
    if not required_ok:
        st.error("⚠️ Colonnes obligatoires introuvables : " + ", ".join(missing_required))

    # ── ÉTAPE 2 : Pièce jointe ──────────────────────────────
    st.subheader("② Pièce jointe email (optionnel)")

    ATTACHMENT_TITUL = resolve_attachment_path(
        "FR EPE - HICHMINE Mohamed Topographe.xlsx",
        name_prefix="FR EPE -",
    )
    ATTACHMENT_PROLONG = resolve_attachment_path(
        "Model PERIODE ESSAI NV.xlsx",
        name_prefix="Model PERIODE ESSAI",
    )

    if is_titularisation:
        default_att_path = ATTACHMENT_TITUL
    else:
        default_att_path = ATTACHMENT_PROLONG

    use_default_att = os.path.exists(default_att_path)
    custom_att = None

    if is_titularisation:
        st.info("La pièce jointe est générée automatiquement pour chaque collaborateur après la section ④ Messages générés.")
    else:
        use_default_att = st.checkbox(
            f"Utiliser **{os.path.basename(default_att_path)}** comme pièce jointe",
            value=use_default_att,
        )
        if not use_default_att:
            custom_att = st.file_uploader(
                "Charger une autre pièce jointe",
                type=["xlsx", "xls", "pdf", "doc", "docx"],
                key="custom_att",
            )

    def get_attachment_for_recipient(nom, prenom, poste=None, direction=None, sup=None, date_entree=None, date_tit=None):
        """Retourne un fichier binaire xlsx personnalisé pour le collaborateur.

        Si le template existe, on le charge, on remplace des valeurs d'exemple
        par les valeurs fournies et on renvoie le fichier modifié en mémoire.
        """
        # cas titularisation : on personnalise à partir du template
        if is_titularisation and os.path.exists(default_att_path):
            try:
                wb = load_workbook(default_att_path)
                replacements = {
                    "HICHMINE": nom or "",
                    "Mohamed": prenom or "",
                    "Topographe": poste or "",
                    "TOARC 4 Tronçon 2": direction or "",
                    "15/09/2025": date_tit.strftime('%d/%m/%Y') if isinstance(date_tit, (datetime,)) else (date_tit or ""),
                    "EL OUANASS Hamza": sup or "",
                }
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                val = cell.value
                                for token, repl in replacements.items():
                                    if token in val and repl is not None:
                                        val = val.replace(token, str(repl))
                                cell.value = val

                bio = _io.BytesIO()
                wb.save(bio)
                bio.seek(0)
                safe_nom = "_".join(nom.split()) if nom else "NOM"
                safe_prenom = "_".join(prenom.split()) if prenom else "PRENOM"
                fname = f"FR_EPE_{safe_nom}_{safe_prenom}.xlsx"
                return bio.read(), fname
            except Exception:
                return None, None

        # cas prolongement ou fichier custom
        if custom_att:
            return custom_att.read(), custom_att.name
        if use_default_att and os.path.exists(default_att_path):
            with open(default_att_path, "rb") as f:
                return f.read(), os.path.basename(default_att_path)
        return None, None

    # ── ÉTAPE 3 : Génération ────────────────────────────────
    st.subheader("③ Générer les messages")

    if required_ok and st.button("🚀 Générer les messages", type="primary", use_container_width=True):
        messages, subjects, errors = [], [], []

        for idx, row in df.iterrows():
            try:
                nom       = get_safe_str(row, col_nom).upper()
                prenom    = get_safe_str(row, col_prenom)
                poste     = get_safe_str(row, col_lib)
                direction = get_safe_str(row, col_lib80)
                sup       = get_safe_str(row, col_sup)
                individu  = get_safe_str(row, col_individu)
                # civilité supprimée — on utilise la valeur par défaut de get_gender_info
                titre, is_f = get_gender_info(None)

                date_entree = parse_date(row[col_date])
                date_renouvellement = parse_date(row[col_date_renouv])
                if date_entree is None:
                    errors.append(f"Ligne {idx + 2} — date invalide pour {nom} {prenom}")
                    messages.append(""); subjects.append("")
                    continue
                if date_renouvellement is None:
                    errors.append(f"Ligne {idx + 2} — Renouvellement Date invalide pour {nom} {prenom}")
                    messages.append(""); subjects.append("")
                    continue

                duree_essai_jours = (date_renouvellement - date_entree).days
                if duree_essai_jours < 0:
                    errors.append(f"Ligne {idx + 2} — Renouvellement Date antérieure à DATE ENTREE pour {nom} {prenom}")
                    messages.append(""); subjects.append("")
                    continue

                date_fin_1ere = date_entree + timedelta(days=duree_essai_jours)
                date_tit = date_entree + timedelta(days=duree_essai_jours)

                if is_titularisation:
                    msg  = build_titularisation(nom, prenom, poste, date_entree, date_tit, titre, is_f)
                    subj = f"Titularisation — {nom} {prenom} — {poste}"
                else:
                    msg  = build_prolongement(individu, nom, prenom, poste, direction, sup,
                                              date_entree, date_fin_1ere, titre, is_f)
                    subj = f"Prolongement Période d'Essai — {nom} {prenom}"

                messages.append(msg)
                subjects.append(subj)

            except Exception as e:
                errors.append(f"Ligne {idx + 2} — {e}")
                messages.append(""); subjects.append("")

        st.session_state["messages"]   = messages
        st.session_state["subjects"]   = subjects
        st.session_state["df_gen"]     = df
        st.session_state["gen_cols"]   = {
            "nom": col_nom, "prenom": col_prenom,
            "poste": col_lib, "direction": col_lib80, "sup": col_sup,
            "date": col_date, "date_renouv": col_date_renouv,
        }
        if errors:
            st.warning("⚠️ Erreurs détectées :\n" + "\n".join(errors))

    # ── ÉTAPE 5 : Résultats ─────────────────────────────────
    if "messages" in st.session_state and st.session_state["messages"]:
        messages    = st.session_state["messages"]
        subjects    = st.session_state["subjects"]
        df_gen      = st.session_state["df_gen"]
        gen_cols    = st.session_state["gen_cols"]

        valid = [(i, m, s)
                 for i, (m, s) in enumerate(zip(messages, subjects))
                 if m]

        if not valid:
            st.error("Aucun message n'a pu être généré. Vérifiez le fichier importé.")
        else:
            st.success(f"✅ **{len(valid)} message(s) généré(s)**")

            # Bouton téléchargement global
            all_text = ""
            for i, msg, subj in valid:
                row = df_gen.iloc[i]
                n = get_safe_str(row, gen_cols["nom"]).upper()
                p = get_safe_str(row, gen_cols["prenom"])
                all_text += f"{'=' * 60}\n{n} {p}\nObjet : {subj}\n{'=' * 60}\n{msg}\n\n"

            filename_dl = (
                f"messages_titularisation_{datetime.now().strftime('%Y%m%d')}.txt"
                if is_titularisation
                else f"messages_prolongement_{datetime.now().strftime('%Y%m%d')}.txt"
            )
            st.download_button(
                "⬇️ Télécharger tous les messages (.txt)",
                data=all_text.encode("utf-8"),
                file_name=filename_dl,
                mime="text/plain",
                use_container_width=True,
            )

            st.divider()
            st.subheader("④ Messages générés")

            for i, msg, subj in valid:
                row = df_gen.iloc[i]
                n = get_safe_str(row, gen_cols["nom"]).upper()
                p = get_safe_str(row, gen_cols["prenom"])
                label = f"👤  {n} {p}"

                with st.expander(label, expanded=False):
                    st.caption(f"**Objet proposé :** {subj}")
                    edited = st.text_area(
                        "✏️ Message (modifiable avant envoi)",
                        value=msg,
                        height=300,
                        key=f"msg_edit_{i}",
                    )

                    if is_titularisation:
                        # récupérer les autres champs depuis le DF pour personnaliser l'attachement
                        poste_val = get_safe_str(row, gen_cols.get("poste"))
                        direction_val = get_safe_str(row, gen_cols.get("direction"))
                        sup_val = get_safe_str(row, gen_cols.get("sup"))
                        date_entree_val = parse_date(row.get(gen_cols.get("date"))) if gen_cols.get("date") else None
                        date_tit_val = parse_date(row.get(gen_cols.get("date_renouv"))) if gen_cols.get("date_renouv") else None
                        att_b_preview, att_n_preview = get_attachment_for_recipient(n, p, poste=poste_val, direction=direction_val, sup=sup_val, date_entree=date_entree_val, date_tit=date_tit_val)
                        if att_b_preview and att_n_preview:
                            st.download_button(
                                "📎 Télécharger la pièce jointe générée",
                                data=att_b_preview,
                                file_name=att_n_preview,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_att_{i}",
                            )
                        else:
                            st.caption("_Pièce jointe de titularisation indisponible pour ce collaborateur._")

else:
    # ── Page d'accueil (aucun fichier) ──────────────────────
    st.info(
        """
        **Comment utiliser cet outil :**

        1. Choisissez le **type de message** dans la barre latérale
        2. **Importez** votre fichier Excel / CSV avec les colonnes collaborateurs
        3. Cliquez sur **Générer les messages**
        4. **Téléchargez** le fichier texte et les pièces jointes générées

        ---
        **Colonnes supportées dans votre fichier :**

        | Colonne | Description | Obligatoire |
        |---------|-------------|-------------|
        | NOM | Nom de famille | ✅ |
        | PRENOM | Prénom | ✅ |
        | LIB / POSTE | Intitulé du poste | ✅ |
        | DATE ENTREE | Date d'entrée en fonction | ✅ |
        | Renouvellement Date | Date de fin d'essai calculée | ✅ |
        | LIB80 / DIRECTION | Direction / Chantier | — |
        | SUP | Nom du responsable hiérarchique | — |
        | INDIVIDU / MAT | Matricule | — |
        | EMAIL | Email du destinataire | — |
        """
    )
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 📄 Exemple — Titularisation")
        st.code(
            "Bonjour ,\n\n"
            "Je me permets de vous contacter au sujet de la\n"
            "titularisation de Mme NOM PRENOM, recrutée en\n"
            "qualité de POSTE depuis le JJ/MM/AAAA.\n"
            "Sa période d'essai arrive à son terme le JJ/MM/AAAA.\n\n"
            "De ce fait, je vous prie de bien vouloir remplir le\n"
            "document ci-joint afin de confirmer ou non sa titularisation.\n\n"
            "Sincères salutations,",
            language=None,
        )
    with col2:
        st.markdown("#### 🔄 Exemple — Prolongement")
        st.code(
            "Bonjour ,\n\n"
            "Nous vous informons que la collaboratrice mentionnée\n"
            "ci-dessous arrive à la fin de leur première période d'essai.\n\n"
            "Mme  NOM  PRENOM  FONCTION  CHANTIER  SUP  DATE  FIN 1ERE\n"
            "MAT  ...  ...     ...       ...        ...  ...   ...\n\n"
            "Pouvez-vous nous confirmer si vous les considérez aptes\n"
            "à bénéficier d'une prolongation de leur période d'essai ?\n"
            "Nous vous prions de bien vouloir nous répondre par retour de mail.\n\n"
            "Cordialement,",
            language=None,
        )

